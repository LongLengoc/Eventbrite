import openpyxl
import json
import smtplib
import email.encoders
import os
from configparser import ConfigParser
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from PyPDF2 import PdfFileReader, PdfFileWriter

class ex2_1():
    def __init__(self):
        #dictionary for save information: email, name, id_ticket, ticket_pdf of each attendee
        self.infor = dict() 
        #get private information
        config = ConfigParser()
        config.read(os.getcwd() + "/config_eventbrite.txt")
        self.account = config.get("Gmail", "Account")
        self.password = config.get("Gmail", "Password")
        self.server = smtplib.SMTP('smtp.gmail.com', 587)
        self.server.starttls()
        self.server.login(self.account, self.password)

    def read_excel(self):
        wb = openpyxl.load_workbook('Danh sach diem danh.xlsx')
        sheet = wb.active
        id_ticket = 11478145771536667379001
        # save information
        # 11: the first line of list attendee in excel
        for i in range (11, sheet.max_row + 1):
            list_email = dict()
            name = sheet.cell(row=i, column=2)
            email = sheet.cell(row=i, column=4)
            list_email['email'] = email.value
            list_email['name'] = name.value
            list_email['id_ticket'] = str(id_ticket)
            # each ticket has ticket code = ticket code before + 2000
            id_ticket += 2000
            # i - 10 : set id attendee -> 1, 2, .... ( because i in range (11, 12, ..))
            self.infor[i - 10] = list_email
        

    def split(self, path):
        #split file pdf ticket into small, each attendee has 1 file ticket pdf
        pdf = PdfFileReader(path)
        for page in range (pdf.getNumPages()):
            pdf_writer = PdfFileWriter()
            pdf_writer.addPage(pdf.getPage(page))
            #set title
            title = 'Ticket_event_' + self.infor[page + 1]['id_ticket']
            output = f'{title}.pdf'
            #save infor
            self.infor[page + 1]['ticket_pdf'] = title + '.pdf'
            with open(output, 'wb') as output_pdf:
                pdf_writer.write(output_pdf)
        print(self.infor)
        
    def send_mail(self):
        try:
            for key in self.infor:
                receiver, text = self.body_maker(self.infor[key])
                self.server.sendmail(self.account, receiver, text)
                self.server.quit
        except Exception as e:
            print(str(e))

    def body_maker(self, receiverDict):
        self.msg = MIMEMultipart()
        self.msg['From'] = self.account
        self.msg['To'] = receiverDict['email']
        self.msg['Subject'] = "Ve check in event " + receiverDict['ticket_pdf']
        text_body = "Kinh gui anh/chi " + receiverDict['name'] + " ve check in event\n Hhhhhh\nhhhhhhh"
        self.body = text_body
        self.filename = receiverDict['ticket_pdf']
        self.msg.attach(MIMEText(self.body, 'plain'))
        attachment = open(self.filename, "rb")
        part = MIMEBase('application', 'octet-stream')
        part.set_payload((attachment).read())
        email.encoders.encode_base64(part)
        part.add_header('Content-Disposition', "attachment; filename = %s" %self.filename)
        self.msg.attach(part)
        text = self.msg.as_string()
        return receiverDict['email'],text

def main():
    ex2 = ex2_1()
    ex2.read_excel()
    ex2.split('81272009769-1147814577-ticket.pdf')
    ex2.send_mail()
    

if __name__ == "__main__":
    main()

